import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl import load_workbook

st.title("HP OID lekérő + link generátor + 300 DPI PNG képek")

uploaded_file = st.file_uploader(
    "Töltsd fel az Excel fájlt (A oszlopban a cikkszámok)",
    type=["xlsx"]
)

if uploaded_file:

    # ===== Excel beolvasás =====
    df = pd.read_excel(uploaded_file)

    # ===== Új oszlopok =====
    df["OID"] = ""
    df["LINK"] = ""
    df["OPEN PRODUCT LINK"] = ""
    df["IMAGE LINK"] = ""
    df["OPEN IMAGE LINK"] = ""

    # ===== API endpointok =====
    base_api = "https://pcb.inc.hp.com/api/catalogs/hu-hu/nodes/search/autocomplete"

    image_api_template = (
        "https://pcb.inc.hp.com/api/catalogs/hu-hu/nodes/{}/contents/I"
        "?status[]=L&status[]=O"
    )

    # ===== Link template-ek =====
    product_link_template = (
        "https://pcb.inc.hp.com/webapp/#/hu-hu/{}/T"
        "?hierarchy=F&status=L&status=O"
    )

    image_link_template = (
        "https://pcb.inc.hp.com/webapp/#/hu-hu/{}/I"
        "?hierarchy=F&status=L&status=O"
    )

    headers = {
        "User-Agent": "Mozilla/5.0"
    }

    progress = st.progress(0)
    total_rows = len(df)

    # ===== Feldolgozás =====
    for i, row in df.iterrows():

        # első oszlop értéke
        prodnum = str(row.iloc[0]).strip()

        # üres sorok kihagyása
        if prodnum == "" or prodnum.lower() == "nan":
            continue

        params = {
            "query": prodnum,
            "status[]": ["L", "O"],
            "exactSearch": "false"
        }

        try:
            response = requests.get(
                base_api,
                params=params,
                headers=headers,
                timeout=10
            )

            if response.status_code == 200:

                data = response.json()

                if "results" in data and len(data["results"]) > 0:

                    # ===== OID =====
                    oid = data["results"][0]["oid"]

                    product_link = product_link_template.format(oid)
                    image_link = image_link_template.format(oid)

                    df.at[i, "OID"] = str(oid)
                    df.at[i, "LINK"] = product_link
                    df.at[i, "IMAGE LINK"] = image_link

                    # ===== Képek lekérése =====
                    try:

                        img_response = requests.get(
                            image_api_template.format(oid),
                            headers=headers,
                            timeout=10
                        )

                        if img_response.status_code == 200:

                            img_data = img_response.json()

                            pic_index = 1

                            # FIGYELEM: contents lista
                            for item in img_data.get("contents", []):

    dpi = item.get("dpiResolution")
    doc_type = item.get("documentTypeDetail")
    image_url = item.get("imageUrlHttps")

    if (
        dpi is not None
        and "300" in str(dpi)
        and doc_type is not None
        and (
            "product image" in str(doc_type).lower()
            or "product image hero" in str(doc_type).lower()
        )
        and image_url
        and image_url.lower().endswith(".png")
    ):
        col_name = f"PIC LINK {pic_index}"
        df.at[i, col_name] = image_url
        pic_index += 1

                    except Exception:
                        pass

                else:
                    df.at[i, "OID"] = "Nincs találat"

            else:
                df.at[i, "OID"] = f"API hiba {response.status_code}"

        except Exception as e:
            df.at[i, "OID"] = str(e)

        progress.progress((i + 1) / total_rows)

    st.success("Feldolgozás kész!")
    st.dataframe(df)

    # ===== Excel export =====
    output = BytesIO()

    df.to_excel(output, index=False)

    output.seek(0)

    # ===== openpyxl betöltés =====
    wb = load_workbook(output)
    ws = wb.active

    # ===== Hyperlinkek =====
    for row_num in range(2, len(df) + 2):

        product_link = ws[f"C{row_num}"].value
        image_link = ws[f"E{row_num}"].value

        # ===== Product OPEN =====
        if product_link:
            ws[f"D{row_num}"].value = "OPEN"
            ws[f"D{row_num}"].hyperlink = product_link
            ws[f"D{row_num}"].style = "Hyperlink"

        # ===== Image OPEN =====
        if image_link:
            ws[f"F{row_num}"].value = "OPEN"
            ws[f"F{row_num}"].hyperlink = image_link
            ws[f"F{row_num}"].style = "Hyperlink"

    # ===== Mentés =====
    final_output = BytesIO()

    wb.save(final_output)

    final_output.seek(0)

    st.download_button(
        label="Excel letöltése",
        data=final_output,
        file_name="output_with_links_and_images.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
